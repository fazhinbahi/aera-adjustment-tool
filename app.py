import streamlit as st
import json
import os
import copy
import zlib
import base64
from datetime import datetime
import io
import pandas as pd

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Constants ────────────────────────────────────────────────────────────────

REASON_CODES = [
    "NPD: New Product/SKU in Market",
    "Distribution in customer(s) (increase or decrease)",
    "Promotion (price; display; VAP; etc)",
    "Sustained change in Price (e.g price increase / duty Increase)",
    "Product lifecycle (end of life)",
    "New Advertising Impact",
    "Market Trend Adjustment",
    "Extreme outliers (e.g. global pandemic)",
    "Changes in regulations",
    "Distributor / Customer on Hold",
    "Other",
]

MESES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

ACTION_TYPES = [
    "Increase/Decrease %",
    "Add or Remove Units",
    "Set to exact number",
    "Remove from month onwards",
    "Remove particular months",
    "Match figures",
    "Other",
]

# ─── Helpers ──────────────────────────────────────────────────────────────────

def get_months_from_cycle(cycle: str) -> list[str]:
    try:
        m = int(cycle.split("-")[1])
    except Exception:
        m = 0
    return MESES[m:]


def load_skus_data(data: dict) -> list[dict]:
    rows = []
    for market in data.get("markets", []):
        for submarket in market.get("submarkets", []):
            subseg = submarket.get("Sub-Segments", "")
            for sku in submarket.get("skus", []):
                country = sku.get("Country Name", "")
                if country == "Not Set":
                    continue
                rows.append({
                    "Sub-Segments": subseg,
                    "Country Name": country,
                    "External Material Group Description": sku.get("External Material Group Description", ""),
                    "Sub-Brand Long Description": sku.get("Sub-Brand Long Description", ""),
                    "Volume": sku.get("Volume", 0.0),
                    "Material Number": sku.get("Material Number", ""),
                })
    return rows


def build_excel_bytes(ajustes: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Adjustments"

    headers = [
        "Cycle", "Sub-Segment", "Country", "Customer",
        "Area Manager", "Regional Director", "Demand Planner",
        "Full Product Path",
        "Category", "Brand", "Sub-brand", "Size", "SKU Number",
        "Period", "Task", "Reason Code", "Current Value", "Date", "Comments",
    ]

    NAVY = "1F3864"; GREEN = "375623"; WHITE = "FFFFFF"
    SOFT_GRN = "F0F7EE"; GOLD = "FFF2CC"; GOLD_HDR = "7F6000"; AMBER_H = "BF8F00"
    BORDER = "BFBFBF"; AMB_FILL = "FFF9E6"

    def hdr(cell, bg, fg=WHITE):
        cell.font = Font(name="Arial", bold=True, color=fg, size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin", color=BORDER), right=Side(style="thin", color=BORDER),
            bottom=Side(style="thin", color=BORDER),
        )

    ws.row_dimensions[1].height = 38
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        if ci == 8:          hdr(cell, GREEN)
        elif 9 <= ci <= 13:  hdr(cell, GOLD_HDR)
        elif ci == 17:       hdr(cell, AMBER_H)
        else:                hdr(cell, NAVY)

    pf     = Font(name="Arial", size=10)
    grn_f  = PatternFill("solid", fgColor=SOFT_GRN)
    gld_f  = PatternFill("solid", fgColor=GOLD)
    alt_f  = PatternFill("solid", fgColor="EBF0FA")
    amb_f  = PatternFill("solid", fgColor=AMB_FILL)

    for ri, a in enumerate(ajustes, 2):
        cat = a.get("Category", ""); brand = a.get("Brand", ""); sb = a.get("Sub-Brand", "")
        vol = a.get("Volume", ""); sku = a.get("SKU", ""); cust = a.get("Customer", "All") or "All"
        parts = [x for x in [cat, brand, sb] if x]
        if vol and vol != "N/A": parts.append(f"{vol}L")
        if sku: parts.append(sku)
        full = " > ".join(parts)
        size = f"{vol}L" if vol and vol != "N/A" else ""

        row_data = [
            a.get("Cycle", ""), a.get("Sub-Segment", ""), a.get("Country", ""),
            cust, a.get("Area Manager", ""), a.get("Regional Director", ""),
            a.get("Demand Planner", ""), full,
            cat, brand, sb, size, sku,
            datetime.now().strftime("%Y"),
            a.get("Task", ""), a.get("Reason Code", ""),
            a.get("Current Value", ""), a.get("Timestamp", "")[:10], a.get("Notes", ""),
        ]
        is_alt = ri % 2 == 0
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(ri, ci, val)
            cell.font = pf
            cell.alignment = Alignment(
                horizontal="center" if ci in (1, 2, 3, 4, 14, 18) else "left",
                vertical="center", wrap_text=(ci in (8, 15, 19)),
            )
            if ci == 8:         cell.fill = grn_f
            elif 9 <= ci <= 13: cell.fill = gld_f
            elif ci == 17:      cell.fill = amb_f
            elif is_alt:        cell.fill = alt_f

    col_widths = [12, 14, 12, 16, 16, 18, 16, 44, 18, 18, 16, 10, 14, 10, 32, 28, 14, 13, 35]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def auto_backup(ajustes: list[dict]):
    try:
        os.makedirs(os.path.join("exports", "backups"), exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join("exports", "backups", f"backup_{ts}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(ajustes, f, indent=2, ensure_ascii=False)
    except Exception:
        pass  # silently ignore on read-only filesystems (e.g. Streamlit Cloud)


def load_from_excel_bytes(file) -> list[dict]:
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, ci).value for ci in range(1, ws.max_column + 1)]
    headers_clean = [str(h).strip() if h else "" for h in headers]
    if "Task" not in headers_clean:
        raise ValueError("Expected column 'Task' not found. Use a file exported by this tool.")
    col_map = {h: i for i, h in enumerate(headers_clean) if h}

    def get(row, key, default=""):
        if key in col_map:
            v = row[col_map[key]]
            return str(v).strip() if v is not None else default
        return default

    loaded = []
    for ri in range(2, ws.max_row + 1):
        row = [ws.cell(ri, ci).value for ci in range(1, len(headers_clean) + 1)]
        if not any(v for v in row if v is not None):
            continue
        task = get(row, "Task")
        if not task:
            continue
        size = get(row, "Size", "")
        vol = size.replace("L", "").strip() if size else ""
        product = get(row, "Full Product Path") or "General"
        loaded.append({
            "Cycle":             get(row, "Cycle"),
            "Area Manager":      get(row, "Area Manager"),
            "Regional Director": get(row, "Regional Director"),
            "Demand Planner":    get(row, "Demand Planner"),
            "Sub-Segment":       get(row, "Sub-Segment"),
            "Country":           get(row, "Country"),
            "Customer":          get(row, "Customer") or "All",
            "Category":          get(row, "Category"),
            "Brand":             get(row, "Brand"),
            "Sub-Brand":         get(row, "Sub-brand"),
            "SKU":               get(row, "SKU Number"),
            "Volume":            vol,
            "Product":           product,
            "Task":              task,
            "Current Value":     get(row, "Current Value"),
            "Reason Code":       get(row, "Reason Code"),
            "Notes":             get(row, "Comments"),
            "Instruction":       task,
            "Timestamp":         get(row, "Date"),
        })
    wb.close()
    return loaded


# ─── Session state initialisation ─────────────────────────────────────────────

def init_state():
    defaults = {
        "ajustes":      [],
        "skus_data":    [],
        "data_loaded":  False,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

# ─── Main app ─────────────────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="Aera Adjustment Assistant v2.3",
        layout="wide",
        page_icon="📊",
    )
    init_state()

    st.title("📊 Aera Adjustment Assistant v2.3")

    # ── Load product catalog ───────────────────────────────────────────────────
    if not st.session_state.data_loaded:
        if "catalog_json" in st.secrets:
            raw = zlib.decompress(base64.b64decode(st.secrets["catalog_json"]))
            st.session_state.skus_data = load_skus_data(json.loads(raw))
            st.session_state.data_loaded = True
        else:
            candidates = [f for f in os.listdir(".") if f.startswith("Adjustments_Forecast") and f.endswith(".json")]
            if candidates:
                archivo = max(candidates, key=os.path.getmtime)
                with open(archivo, "r", encoding="utf-8") as f:
                    st.session_state.skus_data = load_skus_data(json.load(f))
                st.session_state.data_loaded = True
                st.caption(f"Catalog loaded from **{archivo}**")
            else:
                st.warning("No product catalog found. Please contact the administrator.")
                uploaded_catalog = st.file_uploader(
                    "Upload the Adjustments_Forecast JSON catalog file", type="json", key="catalog_upload"
                )
                if uploaded_catalog:
                    st.session_state.skus_data = load_skus_data(json.load(uploaded_catalog))
                    st.session_state.data_loaded = True
                    st.rerun()
                st.stop()

    skus_data: list[dict] = st.session_state.skus_data

    # ── SESSION FILTERS ────────────────────────────────────────────────────────
    with st.expander("⚙️ Session Filters (fixed for this meeting)", expanded=True):
        c1, c2, c3, c4 = st.columns(4)
        cycle          = c1.text_input("Cycle",             value="2026-05", key="f_cycle")
        area_manager   = c2.text_input("Area Manager",      value="JuanMi",  key="f_am")
        regional_dir   = c3.text_input("Regional Director", value="Rafael",  key="f_rd")
        demand_planner = c4.text_input("Demand Planner",    value="Enrique", key="f_dp")

        c5, c6 = st.columns(2)
        subseg_opts = [""] + sorted({s["Sub-Segments"] for s in skus_data})
        subseg = c5.selectbox("Sub-Segment", subseg_opts, key="f_subseg")

        country_opts = [""]
        if subseg:
            country_opts += sorted({s["Country Name"] for s in skus_data if s["Sub-Segments"] == subseg})
        country = c6.selectbox("Country", country_opts, key="f_country")

    months_from_cycle = get_months_from_cycle(cycle)

    # ── NEW ADJUSTMENT ─────────────────────────────────────────────────────────
    st.subheader("New Adjustment")

    ca, cb = st.columns(2)
    category = ca.text_input("Category (optional)", key="adj_category")
    customer = cb.text_input('Customer (optional — blank = "All")', key="adj_customer")

    not_in_catalog = st.checkbox("Product not in catalog (manual entry)", key="adj_nic")

    filtered = [
        s for s in skus_data
        if s["Sub-Segments"] == subseg and s["Country Name"] == country
    ] if subseg and country else []

    # Brand
    cb1, cb2 = st.columns(2)
    if not_in_catalog:
        brand    = cb1.text_input("Brand",     key="adj_brand_man")
        subbrand = cb2.text_input("Sub-Brand", key="adj_subbrand_man")
        sku_number = ""
        volume_val = ""
        cc1, cc2 = st.columns(2)
        sku_number = cc1.text_input("SKU Number", key="adj_sku_man")
        volume_val = cc2.text_input("Volume (L)", key="adj_vol_man")
        sku_display = sku_number
    else:
        brand_opts = [""] + sorted({s["External Material Group Description"] for s in filtered})
        brand = cb1.selectbox("Brand", brand_opts, key="adj_brand")

        filt_sb = [s for s in filtered if not brand or s["External Material Group Description"] == brand]
        sb_opts = [""] + sorted({s["Sub-Brand Long Description"] for s in filt_sb})
        subbrand = cb2.selectbox("Sub-Brand", sb_opts, key="adj_subbrand")

        filt_sku = [
            s for s in filtered
            if (not brand    or s["External Material Group Description"] == brand)
            and (not subbrand or s["Sub-Brand Long Description"] == subbrand)
        ]
        sku_opts    = ["", "New SKU"] + sorted({f"{s['Material Number']} ({s['Volume']}L)" for s in filt_sku})
        cc1, cc2    = st.columns([3, 1])
        sku_display = cc1.selectbox("SKU", sku_opts, key="adj_sku")

        if sku_display == "New SKU":
            cn1, cn2 = st.columns(2)
            sku_number = cn1.text_input("New SKU Number",      key="adj_newsku_num")
            volume_val = cn2.text_input("Volume (L, optional)", key="adj_newsku_vol")
        elif sku_display:
            sku_number = sku_display.split(" (")[0]
            vol_match  = next((s["Volume"] for s in filt_sku if s["Material Number"] == sku_number), "")
            volume_val = str(vol_match) if vol_match else ""
            cc2.markdown(f"<br><b>{volume_val} L</b>", unsafe_allow_html=True)
        else:
            sku_number = ""
            volume_val = ""

    # ── Action type ───────────────────────────────────────────────────────────
    st.markdown("**Action type**")
    action_type = st.selectbox("Action type", [""] + ACTION_TYPES, key="adj_action", label_visibility="collapsed")

    # Params rendered per action
    current_value      = ""
    several_months_keys: list[str] = []  # populated for multi-month actions
    match_months_keys: dict[str, tuple[str, str]] = {}

    if action_type == "Increase/Decrease %":
        p1, p2, p3 = st.columns(3)
        p1.selectbox("Month",           months_from_cycle, key="ap_month_inc")
        p2.text_input("Percentage (+/-)",                  key="ap_pct")
        current_value = p3.text_input("Current Value",     key="ap_curr_inc")

    elif action_type == "Add or Remove Units":
        p1, p2, p3 = st.columns(3)
        p1.selectbox("Month (single)", [""] + months_from_cycle, key="ap_month_add")
        p2.text_input("Units (+/-)",                               key="ap_units")
        current_value = p3.text_input("Current Value",             key="ap_curr_add")

        with st.expander("📅 Configure several months instead"):
            st.caption("Fill Units and/or Current Value for each month you want to adjust. Leave blank to skip.")
            cols = st.columns(3)
            for i, month in enumerate(months_from_cycle):
                with cols[i % 3]:
                    st.markdown(f"**{month}**")
                    k1 = f"sev_add_v1_{month}"; k2 = f"sev_add_v2_{month}"
                    st.text_input("Units (+/-)",    key=k1, label_visibility="collapsed", placeholder="Units (+/-)")
                    st.text_input("Current Value",  key=k2, label_visibility="collapsed", placeholder="Current Value")
                    several_months_keys.append(month)

    elif action_type == "Set to exact number":
        p1, p2, p3 = st.columns(3)
        p1.selectbox("Month (single)", [""] + months_from_cycle, key="ap_month_set")
        p2.text_input("Desired value",                             key="ap_desired")
        current_value = p3.text_input("Current Value",             key="ap_curr_set")

        with st.expander("📅 Configure several months instead"):
            st.caption("Fill Desired Value and/or Current Value for each month. Leave blank to skip.")
            cols = st.columns(3)
            for i, month in enumerate(months_from_cycle):
                with cols[i % 3]:
                    st.markdown(f"**{month}**")
                    k1 = f"sev_set_v1_{month}"; k2 = f"sev_set_v2_{month}"
                    st.text_input("Desired Value",  key=k1, label_visibility="collapsed", placeholder="Desired Value")
                    st.text_input("Current Value",  key=k2, label_visibility="collapsed", placeholder="Current Value")
                    several_months_keys.append(month)

    elif action_type == "Remove from month onwards":
        p1, p2 = st.columns(2)
        p1.selectbox("Start month", months_from_cycle, key="ap_start_month")
        current_value = p2.text_input("Current Value", key="ap_curr_rem")

    elif action_type == "Remove particular months":
        st.markdown("Select months to remove:")
        rm_cols = st.columns(4)
        for i, mes in enumerate(MESES):
            rm_cols[i % 4].checkbox(mes, key=f"ap_rm_{mes}")
        current_value = st.text_input("Current Value (optional)", key="ap_curr_rempart")

    elif action_type == "Match figures":
        mp1, mp2 = st.columns(2)
        mp1.selectbox("Match with", ["Actual Orders", "LE", "3PD", "Other"], key="ap_match_type")
        if st.session_state.get("ap_match_type") == "Other":
            mp2.text_input("Specify", key="ap_match_other")

        st.markdown("---")
        hc = st.columns(3)
        hc[0].markdown("**Month**"); hc[1].markdown("**Current Adj Forecast**"); hc[2].markdown("**Desired Value**")
        for month in months_from_cycle:
            rc = st.columns(3)
            rc[0].markdown(month)
            kc = f"ap_mc_{month}"; kd = f"ap_md_{month}"
            rc[1].text_input("Current", key=kc, label_visibility="collapsed", placeholder="Current")
            rc[2].text_input("Desired", key=kd, label_visibility="collapsed", placeholder="Desired")
            match_months_keys[month] = (kc, kd)

    elif action_type == "Other":
        st.text_area("Custom description", key="ap_custom_text")
        current_value = st.text_input("Current Value", key="ap_curr_other")

    # Reason code + notes
    st.selectbox("Reason Code *", [""] + REASON_CODES, key="adj_reason")
    if st.checkbox("Additional Notes", key="adj_notes_toggle"):
        st.text_area("Notes", key="adj_notes_text")

    # ── Save / Duplicate / Clear ───────────────────────────────────────────────
    btn1, btn2, btn3 = st.columns([1, 1, 4])

    save_clicked = btn1.button("💾 Save Adjustment", type="primary")
    btn2.button("📋 Duplicate Last", key="dup_last_btn")

    if st.session_state.get("dup_last_btn") and st.session_state.ajustes:
        last = st.session_state.ajustes[-1]
        st.info(
            f"Last adjustment: **{last.get('Product','')}** | "
            f"Reason: {last.get('Reason Code','')}  \n"
            "Change the action fields above and click Save."
        )

    if save_clicked:
        reason = st.session_state.get("adj_reason", "")
        if not reason:
            st.error("Reason code is required.")
        elif not action_type:
            st.error("Please select an action type.")
        else:
            # Build product string
            if not_in_catalog:
                m_val = st.session_state.get("adj_brand_man", "")
                sb_val = st.session_state.get("adj_subbrand_man", "")
                sku_f  = st.session_state.get("adj_sku_man", "")
                vol_f  = st.session_state.get("adj_vol_man", "")
            elif sku_display == "New SKU":
                m_val  = brand; sb_val = subbrand
                sku_f  = st.session_state.get("adj_newsku_num", "")
                vol_f  = st.session_state.get("adj_newsku_vol", "")
            else:
                m_val  = brand; sb_val = subbrand; sku_f = sku_number; vol_f = volume_val

            parts = [x for x in [m_val, sb_val] if x]
            if vol_f and vol_f != "N/A": parts.append(f"{vol_f}L")
            if sku_f: parts.append(sku_f)
            prod = " - ".join([m_val, sb_val]).strip(" -") if (m_val or sb_val) else (f"SKU {sku_f}" if sku_f else "General adjustment")

            notes = st.session_state.get("adj_notes_text", "") if st.session_state.get("adj_notes_toggle") else ""

            base = {
                "Cycle":             cycle,
                "Area Manager":      area_manager,
                "Regional Director": regional_dir,
                "Demand Planner":    demand_planner,
                "Sub-Segment":       subseg,
                "Country":           country,
                "Customer":          st.session_state.get("adj_customer", "").strip() or "All",
                "Category":          st.session_state.get("adj_category", "").strip(),
                "Brand":             m_val,
                "Sub-Brand":         sb_val,
                "SKU":               sku_f,
                "Volume":            vol_f,
                "Product":           prod,
                "Task":              "",
                "Current Value":     "",
                "Reason Code":       reason,
                "Notes":             notes,
                "Instruction":       "",
                "Timestamp":         "",
            }

            saved = 0
            errors = []

            # Multi-month: Add/Remove or Set to exact
            if action_type in ("Add or Remove Units", "Set to exact number") and several_months_keys:
                prefix = "sev_add" if action_type == "Add or Remove Units" else "sev_set"
                for month in several_months_keys:
                    v1 = st.session_state.get(f"{prefix}_v1_{month}", "").strip()
                    v2 = st.session_state.get(f"{prefix}_v2_{month}", "").strip()
                    if not v1 and not v2:
                        continue
                    if not v1:
                        continue
                    a = copy.deepcopy(base)
                    a["Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    a["Task"] = (
                        f"Add/Remove {v1} units in {month}"
                        if action_type == "Add or Remove Units"
                        else f"Set to {v1} in {month}"
                    )
                    a["Current Value"] = v2
                    a["Instruction"] = a["Task"]
                    st.session_state.ajustes.append(a)
                    saved += 1

            # Match figures
            if action_type == "Match figures":
                mt = st.session_state.get("ap_match_type", "Actual Orders")
                if mt == "Other":
                    mt = st.session_state.get("ap_match_other", "") or "Other"
                for month, (kc, kd) in match_months_keys.items():
                    curr_m = st.session_state.get(kc, "").strip()
                    des_m  = st.session_state.get(kd, "").strip()
                    if not curr_m and not des_m:
                        continue
                    a = copy.deepcopy(base)
                    a["Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    a["Task"] = f"Match to {mt} in {month}: desired={des_m}"
                    a["Current Value"] = curr_m
                    a["Instruction"] = a["Task"]
                    st.session_state.ajustes.append(a)
                    saved += 1

            # Single-row actions
            if action_type not in ("Match figures",) and not (
                action_type in ("Add or Remove Units", "Set to exact number") and saved > 0
            ):
                task = ""
                curr = current_value

                if action_type == "Increase/Decrease %":
                    m  = st.session_state.get("ap_month_inc", "")
                    p  = st.session_state.get("ap_pct", "").strip()
                    curr = st.session_state.get("ap_curr_inc", "").strip()
                    task = f"Increase/Decrease {p}% in {m}" if m and p else ""

                elif action_type == "Add or Remove Units":
                    m  = st.session_state.get("ap_month_add", "")
                    u  = st.session_state.get("ap_units", "").strip()
                    curr = st.session_state.get("ap_curr_add", "").strip()
                    task = f"Add/Remove {u} units in {m}" if m and u else ""

                elif action_type == "Set to exact number":
                    m  = st.session_state.get("ap_month_set", "")
                    d  = st.session_state.get("ap_desired", "").strip()
                    curr = st.session_state.get("ap_curr_set", "").strip()
                    task = f"Set to {d} in {m}" if m and d else ""

                elif action_type == "Remove from month onwards":
                    m  = st.session_state.get("ap_start_month", "")
                    curr = st.session_state.get("ap_curr_rem", "").strip()
                    task = f"Remove volume from {m} onwards" if m else ""

                elif action_type == "Remove particular months":
                    sel = [mes for mes in MESES if st.session_state.get(f"ap_rm_{mes}")]
                    curr = st.session_state.get("ap_curr_rempart", "").strip()
                    task = f"Remove volume in {', '.join(sel)}" if sel else ""

                elif action_type == "Other":
                    ct = st.session_state.get("ap_custom_text", "").strip()
                    curr = st.session_state.get("ap_curr_other", "").strip()
                    task = (f"Other: {ct[:40]}..." if len(ct) > 40 else f"Other: {ct}") if ct else ""

                if task:
                    a = copy.deepcopy(base)
                    a["Timestamp"]     = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    a["Task"]          = task
                    a["Current Value"] = curr
                    a["Instruction"]   = task
                    st.session_state.ajustes.append(a)
                    saved += 1
                elif saved == 0:
                    errors.append("Please complete the action parameters.")

            if saved > 0:
                auto_backup(st.session_state.ajustes)
                st.success(f"{saved} adjustment(s) saved!")
                st.rerun()
            for e in errors:
                st.error(e)

    st.divider()

    # ── SESSION TABLE ──────────────────────────────────────────────────────────
    n = len(st.session_state.ajustes)
    st.subheader(f"Session Adjustments ({n})")

    if st.session_state.ajustes:
        df = pd.DataFrame([
            {
                "#":           i + 1,
                "Customer":    a.get("Customer", "All"),
                "Product":     a.get("Product", ""),
                "Task":        a.get("Task", ""),
                "Reason Code": a.get("Reason Code", ""),
                "Timestamp":   a.get("Timestamp", ""),
            }
            for i, a in enumerate(st.session_state.ajustes)
        ])
        st.dataframe(df, use_container_width=True, hide_index=True)

        ta1, ta2, ta3, ta4 = st.columns([2, 1, 1, 1])
        sel_row = ta1.number_input("Select row # to act on", min_value=1, max_value=n, value=1, key="sel_row") - 1

        if ta2.button("🗑️ Delete"):
            del st.session_state.ajustes[sel_row]
            st.rerun()

        if ta3.button("📋 Duplicate"):
            a = copy.deepcopy(st.session_state.ajustes[sel_row])
            a["Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            st.session_state.ajustes.append(a)
            auto_backup(st.session_state.ajustes)
            st.rerun()

        if ta4.button("🗑️ Clear All"):
            auto_backup(st.session_state.ajustes)
            st.session_state.ajustes = []
            st.rerun()
    else:
        st.info("No adjustments yet. Use the form above to add some.")

    st.divider()

    # ── EXPORT / IMPORT ────────────────────────────────────────────────────────
    st.subheader("Export / Import")
    ie1, ie2, ie3 = st.columns(3)

    with ie1:
        st.markdown("**Export to Excel**")
        if st.session_state.ajustes:
            subseg_name  = subseg  or "NoSubSeg"
            country_name = country or "NoCountry"
            ts           = datetime.now().strftime("%Y%m%d_%H%M%S")
            fname        = f"{subseg_name}_{country_name}_PreAlignment_Adjustments_{ts}.xlsx"
            xlsx_bytes   = build_excel_bytes(st.session_state.ajustes)
            st.download_button(
                "📤 Download Excel",
                data=xlsx_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.caption("No adjustments to export yet.")

    with ie2:
        st.markdown("**Load from Excel**")
        up_xlsx = st.file_uploader("Upload exported .xlsx", type="xlsx", key="up_xlsx")
        if up_xlsx:
            try:
                loaded = load_from_excel_bytes(up_xlsx)
                if loaded:
                    st.session_state.ajustes = loaded
                    auto_backup(loaded)
                    st.success(f"Loaded {len(loaded)} adjustment(s).")
                    st.rerun()
                else:
                    st.warning("No valid rows found in the file.")
            except Exception as e:
                st.error(f"Could not load file: {e}")

    with ie3:
        st.markdown("**Load Backup (JSON)**")
        up_json = st.file_uploader("Upload backup .json", type="json", key="up_json")
        if up_json:
            try:
                loaded = json.load(up_json)
                st.session_state.ajustes = loaded
                st.success(f"Loaded {len(loaded)} adjustment(s) from backup.")
                st.rerun()
            except Exception as e:
                st.error(f"Could not load backup: {e}")


if __name__ == "__main__":
    main()
